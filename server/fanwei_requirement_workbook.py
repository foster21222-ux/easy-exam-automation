#!/usr/bin/env python3

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_CONTENT_ITEMS = ("考前等待提示", "欢迎语", "考试承诺书内容")


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def read_requirement_defaults(template_path):
    wb = load_workbook(Path(template_path), read_only=True, data_only=True)
    ws = wb["业务需求单"]
    defaults = {"欢迎语": ""}
    for row in ws.iter_rows(min_row=5):
        if len(row) < 4:
            continue
        item = normalize_text(row[2].value)
        if item not in DEFAULT_CONTENT_ITEMS:
            continue
        value = normalize_text(row[3].value)
        if value:
            defaults[item] = value
    wb.close()
    return defaults


def fill_requirement_workbook(template_path, output_path, fields):
    template = Path(template_path)
    output = Path(output_path)
    wb = load_workbook(template)
    ws = wb["业务需求单"]

    for row in ws.iter_rows(min_row=5):
        if len(row) < 4:
            continue
        item = normalize_text(row[2].value)
        if not item:
            continue
        value = None
        if item in fields:
            value = fields[item]
        else:
            for key, candidate in fields.items():
                if key and key in item:
                    value = candidate
                    break
        if value is None:
            continue
        row[3].value = value

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    if len(sys.argv) == 3 and sys.argv[1] == "--defaults":
        print(json.dumps(read_requirement_defaults(sys.argv[2]), ensure_ascii=False))
        return
    if len(sys.argv) != 4:
        raise SystemExit(
            "Usage: fanwei_requirement_workbook.py <template.xlsx> <payload.json> <output.xlsx>\n"
            "   or: fanwei_requirement_workbook.py --defaults <template.xlsx>"
        )
    template_path = Path(sys.argv[1])
    payload_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    fill_requirement_workbook(template_path, output_path, payload.get("requirementFields") or payload.get("fields") or {})
    print(json.dumps({"outputPath": str(output_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
