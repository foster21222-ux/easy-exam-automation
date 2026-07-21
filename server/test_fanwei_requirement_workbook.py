import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
GENERATOR = ROOT / "server" / "fanwei_requirement_workbook.py"
PARSER = ROOT / "server" / "exam_request_parser.py"
TEMPLATE = ROOT / "template" / "v2易考新建考试需求单.xlsx"


class FanweiRequirementWorkbookTest(unittest.TestCase):
    def cell_value_for_item(self, ws, item_name):
        for row in ws.iter_rows(min_row=5):
            if len(row) >= 4 and str(row[2].value or "").strip() == item_name:
                return row[3].value
        self.fail(f"未找到配置项：{item_name}")

    def test_reads_default_rich_content_and_welcome_text(self):
        defaults = json.loads(
            subprocess.check_output(
                [sys.executable, str(GENERATOR), "--defaults", str(TEMPLATE)],
                text=True,
            )
        )
        self.assertIn("准考证号为个人手机号", defaults["考前等待提示"])
        self.assertIn("color: red", defaults["考前等待提示"])
        self.assertIn("<strong>", defaults["考前等待提示"])
        self.assertEqual(defaults["欢迎语"], "")
        self.assertIn("考试承诺书", defaults["考试承诺书内容"])
        self.assertIn("color: #ff0000", defaults["考试承诺书内容"])
        self.assertIn("<strong>", defaults["考试承诺书内容"])

    def test_generates_parser_ready_workbook_while_preserving_template_defaults(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            payload = {
                "requirementFields": {
                    "考试名称": "四川省通川工程技术开发有限公司校招笔试",
                    "考试日期时间": "2026/7/5 9:30:00-2026/7/5 11:30:00",
                    "试考日期时间": "2026/7/4 10:00:00-2026/7/4 17:00:00",
                    "提前登录时间": "30分钟",
                    "限制迟到时间": "20分钟",
                    "试卷扣时规则": "迟到及离开扣时",
                    "考试地址": "统一考试地址",
                    "视频监控": "需要",
                    "视频录制": "开启录制",
                    "鹰眼监控": "需要",
                    "考试类型": "客户端考试",
                    "登陆次数": "10",
                    "人工判分": "旧版判分（包含系统判分及悦评对接）",
                    "科目信息": "四川省通川工程技术开发有限公司校招笔试",
                }
            }
            payload_path = tmp_path / "payload.json"
            output_path = tmp_path / "fanwei.xlsx"
            payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            subprocess.check_call([sys.executable, str(GENERATOR), str(TEMPLATE), str(payload_path), str(output_path)])

            generated = load_workbook(output_path)
            template = load_workbook(TEMPLATE)
            ws = generated["业务需求单"]
            template_ws = template["业务需求单"]
            self.assertEqual(ws["D5"].value, "四川省通川工程技术开发有限公司校招笔试")
            self.assertEqual(ws["D21"].value, "旧版判分（包含系统判分及悦评对接）")
            self.assertEqual(ws["D10"].value, template_ws["D10"].value)
            self.assertEqual(ws["D14"].value, template_ws["D14"].value)
            self.assertGreaterEqual(len(ws.data_validations.dataValidation), 1)

            parsed = json.loads(subprocess.check_output([sys.executable, str(PARSER), str(output_path)], text=True))
            self.assertEqual(parsed["config"]["examName"], "四川省通川工程技术开发有限公司校招笔试")
            self.assertEqual(parsed["config"]["startTimeDisplay"], "2026/07/05 09:30")
            self.assertEqual(parsed["config"]["mockStartTimeDisplay"], "2026/07/04 10:00")
            self.assertTrue(parsed["config"]["manualScore"])
            self.assertEqual(parsed["config"]["manualScoreText"], "旧版判分（包含系统判分及悦评对接）")
            self.assertTrue(parsed["config"]["clientExam"])
            self.assertEqual(parsed["config"]["clientLoginLimit"], 10)

    def test_blank_rich_fields_override_template_defaults(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            payload = {
                "requirementFields": {
                    "考试名称": "空富文本测试",
                    "考试日期时间": "2026/7/5 9:30:00-2026/7/5 11:30:00",
                    "考前等待提示": "",
                    "考试承诺书内容": "",
                    "科目信息": "综合能力",
                }
            }
            payload_path = tmp_path / "payload.json"
            output_path = tmp_path / "fanwei.xlsx"
            payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            subprocess.check_call([sys.executable, str(GENERATOR), str(TEMPLATE), str(payload_path), str(output_path)])

            generated = load_workbook(output_path)
            ws = generated["业务需求单"]
            self.assertIn(self.cell_value_for_item(ws, "考前等待提示"), (None, ""))
            self.assertIn(self.cell_value_for_item(ws, "考试承诺书内容"), (None, ""))

            parsed = json.loads(subprocess.check_output([sys.executable, str(PARSER), str(output_path)], text=True))
            self.assertEqual(parsed["config"]["preLoginPrompt"], "")
            self.assertEqual(parsed["config"]["pledgeContent"], "")


if __name__ == "__main__":
    unittest.main()
