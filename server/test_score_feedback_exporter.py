import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from score_feedback_exporter import export_score_feedback


class ScoreFeedbackExporterTest(unittest.TestCase):
    def test_exports_scores_from_template_with_text_identifiers(self):
        root = Path(__file__).resolve().parents[1]
        template = root / "template" / "成绩单模板.xlsx"
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            payload = temp / "payload.json"
            output = temp / "score.xlsx"
            rows = []
            for index in range(6):
                rows.append(
                    {
                        "name": f"考生{index + 1}",
                        "gender": "男" if index % 2 == 0 else "女",
                        "identity_id": f"00123456789012345{index}",
                        "mobile": f"0138000000{index:02d}",
                        "email": "384564539@qq.com" if index == 0 else f"user{index}@example.com",
                        "course": "四川省通川工程技术开发有限公司校招笔试" if index == 0 else "综合能力",
                        "permit": f"000{index + 1}",
                        "exam_status": "已完成" if index == 0 else ("未开考" if index == 1 else "异常状态"),
                        "score": "" if index == 1 else 80 + index,
                        "violation": "" if index != 2 else "作弊",
                    }
                )
            payload.write_text(
                json.dumps(
                    {
                        "examName": "客户招聘笔试",
                        "examTime": "2026-07-25 09:00 ~ 2026-07-25 10:30",
                        "processedDate": "2026年6月29日",
                        "rows": rows,
                    },
                    ensure_ascii=False,
                ),
                "utf-8",
            )

            result = export_score_feedback(template, payload, output)

            self.assertTrue(result["ok"])
            workbook = load_workbook(output)
            sheet = workbook.active
            merged_ranges = {str(cell_range) for cell_range in sheet.merged_cells.ranges}
            self.assertTrue({
                "A2:A3",
                "B2:C3",
                "D2:E3",
                "F2:F3",
                "G2:H3",
                "I2:J3",
                "B4:C4",
                "D4:E4",
                "G4:H4",
                "I4:J4",
            }.issubset(merged_ranges))
            self.assertEqual(sheet["B2"].value, "考试名称")
            self.assertEqual(sheet["D2"].value, "考试时间")
            self.assertEqual(sheet["F2"].value, "采购方联系人")
            self.assertEqual(sheet["G2"].value, "项目经理（ATA）")
            self.assertEqual(sheet["I2"].value, "测评经理（ATA）")
            self.assertEqual(sheet["B4"].value, "客户招聘笔试")
            self.assertEqual(sheet["D4"].value, "2026-07-25 09:00 ~ 2026-07-25 10:30")
            self.assertEqual(sheet["G4"].value, "王历平")
            self.assertEqual(sheet["I4"].value, "乔晓钢")
            exam_name_width = sheet.column_dimensions["B"].width + sheet.column_dimensions["C"].width
            exam_time_width = sheet.column_dimensions["D"].width + sheet.column_dimensions["E"].width
            self.assertGreaterEqual(exam_name_width, 42)
            self.assertGreaterEqual(exam_time_width, 34)
            self.assertLessEqual(exam_time_width, 38)
            top_content_cells = [sheet[coordinate] for coordinate in ("A4", "B4", "D4", "F4", "G4", "I4")]
            self.assertTrue(all(cell.alignment.horizontal == "center" for cell in top_content_cells))
            self.assertTrue(all(cell.alignment.vertical == "center" for cell in top_content_cells))
            self.assertTrue(all(cell.alignment.wrap_text for cell in top_content_cells))
            self.assertTrue(all(not cell.alignment.shrink_to_fit for cell in top_content_cells))
            self.assertGreaterEqual(sheet.row_dimensions[4].height, 30)
            self.assertLessEqual(sheet.row_dimensions[4].height, 36)
            self.assertEqual(sheet.page_setup.orientation, "landscape")
            self.assertEqual(sheet.page_setup.fitToWidth, 1)
            self.assertEqual(sheet.column_dimensions["B"].width, 12)
            self.assertEqual(sheet.column_dimensions["C"].width, 30)
            self.assertEqual(sheet.column_dimensions["D"].width, 14)
            self.assertEqual(sheet.column_dimensions["E"].width, 22)
            self.assertEqual(sheet.column_dimensions["F"].width, 44)
            self.assertEqual(sheet.column_dimensions["G"].width, 16)
            self.assertEqual(sheet.column_dimensions["H"].width, 8)
            self.assertEqual(sheet.column_dimensions["I"].width, 8)
            self.assertEqual(sheet.column_dimensions["J"].width, 16)
            self.assertEqual(sheet["G4"].border.left.style, "thin")
            self.assertEqual(sheet["G4"].border.top.style, "thin")
            self.assertEqual(sheet["G4"].border.bottom.style, "thin")
            self.assertEqual(sheet["H4"].border.right.style, "thin")
            self.assertEqual(sheet["I4"].border.left.style, "thin")
            self.assertEqual(sheet["I4"].border.top.style, "thin")
            self.assertEqual(sheet["I4"].border.bottom.style, "thin")
            self.assertEqual(sheet["J4"].border.right.style, "thin")
            report_font_cells = [sheet[coordinate] for coordinate in ("A2", "B2", "D2", "F2", "G2", "I2", "A5")]
            report_font_cells.extend(sheet.cell(6, column) for column in range(1, 11))
            report_font_cells.extend(
                sheet[coordinate]
                for coordinate in ("A4", "B4", "D4", "F4", "G4", "I4")
            )
            report_font_cells.extend(
                sheet.cell(row, column)
                for row in range(7, 7 + len(rows))
                for column in range(1, 11)
            )
            self.assertTrue(all(
                cell.font.name == "微软雅黑" and cell.font.sz == 10
                for cell in report_font_cells
            ))
            gray_header_cells = [sheet[coordinate] for coordinate in ("A2", "B2", "D2", "F2", "G2", "I2", "A5")]
            self.assertTrue(all(
                cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == "FFCCCCCC"
                for cell in gray_header_cells
            ))
            self.assertEqual(sheet["A7"].value, "考生1")
            self.assertEqual(sheet["H7"].value, "参考")
            self.assertEqual(sheet["H8"].value, "缺考")
            self.assertEqual(sheet["J7"].value, "无")
            self.assertEqual(sheet["C7"].number_format, "@")
            self.assertEqual(sheet["D7"].number_format, "@")
            self.assertEqual(sheet["G7"].number_format, "@")
            self.assertEqual(sheet["C7"].value, "001234567890123450")
            self.assertEqual(sheet["D7"].value, "013800000000")
            self.assertEqual(sheet["E7"].value, "384564539@qq.com")
            self.assertEqual(sheet["F7"].value, "四川省通川工程技术开发有限公司校招笔试")
            data_cells = [sheet.cell(row, column) for row in range(7, 7 + len(rows)) for column in range(1, 11)]
            self.assertTrue(all(cell.alignment.horizontal == "center" for cell in data_cells))
            self.assertTrue(all(cell.alignment.vertical == "center" for cell in data_cells))
            self.assertTrue(all(cell.alignment.wrap_text for cell in data_cells))
            self.assertTrue(all(not cell.alignment.shrink_to_fit for cell in data_cells))
            self.assertGreaterEqual(sheet.row_dimensions[7].height, 30)
            self.assertIn("A5:J5", merged_ranges)
            self.assertEqual(sheet["A5"].value, "成绩明细")
            self.assertEqual(sheet["A5"].alignment.horizontal, "center")
            self.assertEqual(sheet["A5"].alignment.vertical, "center")
            yellow_fills = [
                cell.coordinate
                for row in sheet.iter_rows()
                for cell in row
                if cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == "FFFFFF00"
            ]
            self.assertEqual(yellow_fills, [])
            footer_rows = [
                row[0].row
                for row in sheet.iter_rows()
                for cell in row
                if cell.value == "全美在线（北京）科技股份有限公司"
            ]
            self.assertTrue(footer_rows)
            self.assertGreater(footer_rows[0], 11)
            blank_before_footer = [
                sheet.cell(footer_rows[0] - 1, column).value
                for column in range(1, 11)
            ]
            self.assertEqual(blank_before_footer, [None] * 10)
            self.assertEqual(sheet.cell(footer_rows[0] + 1, 6).value, "2026年6月29日")

    def test_exports_valid_without_score_as_absent_with_placeholder(self):
        root = Path(__file__).resolve().parents[1]
        template = root / "template" / "成绩单模板.xlsx"
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            payload = temp / "payload.json"
            output = temp / "score.xlsx"
            payload.write_text(
                json.dumps(
                    {
                        "examName": "缺考状态测试",
                        "examTime": "2026-06-25 19:00 ~ 2026-06-25 21:00",
                        "rows": [
                            {
                                "name": "李科",
                                "permit": "13208164907",
                                "exam_status": "valid",
                                "score": "",
                                "violation": "",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                "utf-8",
            )

            result = export_score_feedback(template, payload, output)

            self.assertTrue(result["ok"])
            sheet = load_workbook(output).active
            self.assertEqual(sheet["H7"].value, "缺考")
            self.assertEqual(sheet["I7"].value, "--")

    def test_appends_assessment_report_hyperlinks_when_present(self):
        root = Path(__file__).resolve().parents[1]
        template = root / "template" / "成绩单模板.xlsx"
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            payload = temp / "payload.json"
            output = temp / "score-with-reports.xlsx"
            payload.write_text(
                json.dumps(
                    {
                        "examName": "测评报告链接测试",
                        "examTime": "2026-07-15 10:00 ~ 2026-07-15 12:00",
                        "processedDate": "2026年7月15日",
                        "rows": [
                            {
                                "name": "刘珈辰",
                                "identity_id": "511381199904240258",
                                "mobile": "13419110769",
                                "email": "liujiachen@example.com",
                                "permit": "13419110769",
                                "exam_status": "已完成",
                                "score": 90,
                                "reports": [
                                    {
                                        "name": "全方位胜任力报告-UCF",
                                        "url": "https://cdn.eztest.cn/report/ucf.pdf",
                                        "status": "finished",
                                    },
                                    {
                                        "name": "情绪倾向报告（标准）-SHLEmotion",
                                        "url": "https://cdn.eztest.cn/report/emotion.pdf",
                                        "status": "finished",
                                    },
                                ],
                            },
                            {
                                "name": "周萌",
                                "identity_id": "510107199808301262",
                                "mobile": "13684016972",
                                "email": "zhoumeng@example.com",
                                "permit": "13684016972",
                                "exam_status": "未开考",
                                "score": "",
                            },
                        ],
                    },
                    ensure_ascii=False,
                ),
                "utf-8",
            )

            result = export_score_feedback(template, payload, output)

            self.assertTrue(result["ok"])
            self.assertEqual(result["reportColumns"], 2)
            sheet = load_workbook(output).active
            merged_ranges = {str(cell_range) for cell_range in sheet.merged_cells.ranges}
            self.assertIn("A5:L5", merged_ranges)
            self.assertEqual(sheet["K6"].value, "全方位胜任力报告-UCF")
            self.assertEqual(sheet["L6"].value, "情绪倾向报告（标准）-SHLEmotion")
            self.assertEqual(sheet["K7"].value, "全方位胜任力报告-UCF")
            self.assertEqual(sheet["L7"].value, "情绪倾向报告（标准）-SHLEmotion")
            self.assertEqual(sheet["K7"].hyperlink.target, "https://cdn.eztest.cn/report/ucf.pdf")
            self.assertEqual(sheet["L7"].hyperlink.target, "https://cdn.eztest.cn/report/emotion.pdf")
            self.assertIsNone(sheet["K8"].value)
            self.assertIsNone(sheet["L8"].value)
            self.assertEqual(sheet["K7"].font.color.rgb, "FF0563C1")
            self.assertEqual(sheet["K7"].font.underline, "single")


if __name__ == "__main__":
    unittest.main()
