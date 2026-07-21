import json
import subprocess
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent.parent
PARSER = ROOT / "server" / "exam_request_parser.py"


class ExamRequestParserTest(unittest.TestCase):
    def parse_workbook(self, rows):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "需求单.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "业务需求单"
            ws.append(["易考新建考试需求单"])
            ws.append(["业务方只需填写“填写内容”列。"])
            ws.append([])
            ws.append(["阶段", "序号", "配置项", "填写内容"])
            for row in rows:
                ws.append(row)
            wb.save(path)
            output = subprocess.check_output([sys.executable, str(PARSER), str(path)], text=True)
            return json.loads(output)

    def test_reads_manual_scoring_when_text_is_present(self):
        result = self.parse_workbook(
            [
                ["基本信息", 1, "考试名称", "新模板测试"],
                ["", 2, "考试日期时间", "2026/7/3 09:00-2026/7/3 10:30"],
                ["", "", "试考日期时间", "2026/7/2 15:00-2026/7/2 20:00"],
                ["考试配置", 16, "人工判分", "旧版判分（包含系统判分及悦评对接）"],
                ["试卷信息", 17, "科目信息", "测试1，测试2"],
            ],
        )

        self.assertTrue(result["config"]["manualScore"])
        self.assertEqual(result["config"]["manualScoreText"], "旧版判分（包含系统判分及悦评对接）")
        self.assertIn(["考试后", "人工判分", "旧版判分（包含系统判分及悦评对接）", "按需求单配置"], result["previewRows"])

    def test_reads_short_and_chinese_time_ranges(self):
        year = datetime.now().year
        result = self.parse_workbook(
            [
                ["基本信息", 1, "考试名称", "短时间格式测试"],
                ["", 2, "考试日期时间", "7-21 15 ：00-16:30"],
                ["", "", "试考日期时间", "7-21 15 点-16 点半"],
                ["试卷信息", 17, "科目信息", "综合能力"],
            ],
        )

        self.assertEqual(result["config"]["startTimeDisplay"], f"{year}/07/21 15:00")
        self.assertEqual(result["config"]["endTimeDisplay"], f"{year}/07/21 16:30")
        self.assertEqual(result["config"]["mockStartTimeDisplay"], f"{year}/07/21 15:00")
        self.assertEqual(result["config"]["mockEndTimeDisplay"], f"{year}/07/21 16:30")

    def test_leaves_manual_scoring_disabled_when_explicitly_not_needed(self):
        result = self.parse_workbook(
            [
                ["基本信息", 1, "考试名称", "仅测评"],
                ["", 2, "考试日期时间", "2026/7/3 09:00-2026/7/3 10:30"],
                ["考试配置", 16, "人工判分", "不需要"],
            ],
        )

        self.assertFalse(result["config"]["manualScore"])

    def test_defaults_watermark_and_copy_protection_to_enabled(self):
        default_result = self.parse_workbook(
            [
                ["基本信息", 1, "考试名称", "默认防护测试"],
                ["", 2, "考试日期时间", "2026/7/3 09:00-2026/7/3 10:30"],
            ],
        )
        disabled_result = self.parse_workbook(
            [
                ["基本信息", 1, "考试名称", "关闭防护测试"],
                ["", 2, "考试日期时间", "2026/7/3 09:00-2026/7/3 10:30"],
                ["考试配置", 20, "答题水印", "否"],
                ["", 21, "禁止复制", "否"],
            ],
        )

        self.assertTrue(default_result["config"]["watermark"])
        self.assertTrue(default_result["config"]["disableCopy"])
        self.assertIn(["考试中", "答题水印", "是", "自动勾选"], default_result["previewRows"])
        self.assertIn(["考试中", "禁止复制", "是", "自动勾选"], default_result["previewRows"])
        self.assertFalse(disabled_result["config"]["watermark"])
        self.assertFalse(disabled_result["config"]["disableCopy"])

    def test_reads_per_course_paper_names_from_subject_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "需求单.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "业务需求单"
            ws.append(["易考新建考试需求单"])
            ws.append(["业务方只需填写“填写内容”列。"])
            ws.append([])
            ws.append(["阶段", "序号", "配置项", "填写内容"])
            ws.append(["基本信息", 1, "考试名称", "每科试卷测试"])
            ws.append(["", 2, "考试日期时间", "2026/7/18 09:00-2026/7/18 10:30"])
            subject = wb.create_sheet("科目信息")
            subject.append(["序号", "科目名称", "科目编号", "试卷名称"])
            subject.append([1, "综合一", "20260718-01-01", "会计学与财务分析基础"])
            subject.append([2, "综合二", "20260718-01-02", "Python语言基础+大数据技术"])
            wb.save(path)

            output = subprocess.check_output([sys.executable, str(PARSER), str(path)], text=True)
            result = json.loads(output)

        self.assertEqual(
            result["config"]["courses"],
            [
                {
                    "name": "综合一",
                    "code": "20260718-01-01",
                    "form_codes": [],
                    "paper_name": "会计学与财务分析基础",
                },
                {
                    "name": "综合二",
                    "code": "20260718-01-02",
                    "form_codes": [],
                    "paper_name": "Python语言基础+大数据技术",
                },
            ],
        )


if __name__ == "__main__":
    unittest.main()
