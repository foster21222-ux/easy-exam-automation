#!/usr/bin/env python3
import copy
import json
import sys
from datetime import date, datetime
from pathlib import Path


DATA_START_ROW = 7
HEADERS = ("姓名", "性别", "证件号码", "手机号码", "邮箱", "科目", "准考证号", "考试状态", "得分", "违纪情况")
BASE_COLUMN_COUNT = len(HEADERS)
TEXT_COLUMNS = (3, 4, 7)
YELLOW_FILL_RGB = "FFFFFF00"
TOP_HEADER_MERGES = (
    "A2:A3",
    "B2:C3",
    "D2:E3",
    "F2:F3",
    "G2:H3",
    "I2:J3",
    "B4:C4",
    "D4:E4",
    "G4:H4",
    "I4:J4",
)
TOP_HEADER_COLUMN_WIDTHS = {
    "B": 12.0,
    "C": 30.0,
    "D": 14.0,
    "E": 22.0,
    "F": 44.0,
    "G": 16.0,
    "H": 8.0,
    "I": 8.0,
    "J": 16.0,
}
HEADER_FILL_RGB = "FFCCCCCC"
BODY_FONT_NAME = "微软雅黑"
BODY_FONT_SIZE = 10
CONTENT_ROW_HEIGHT = 30
TOP_CONTENT_ROW_HEIGHT = 34


def text(value):
    if value is None:
        return ""
    return str(value)


def normalized_status(value):
    raw = text(value).strip()
    if not raw:
        return "缺考"
    if raw == "已完成":
        return "参考"
    if raw == "未开考" or raw.lower() == "valid":
        return "缺考"
    return raw


def score_value(value):
    if value is None or text(value).strip() == "":
        return ""
    raw = text(value).strip()
    try:
        numeric = float(raw)
        return int(numeric) if numeric.is_integer() else numeric
    except ValueError:
        return raw


def find_footer_row(sheet):
    for row in range(DATA_START_ROW, sheet.max_row + 1):
        values = [text(sheet.cell(row, column).value) for column in range(1, sheet.max_column + 1)]
        joined = "".join(values)
        if "全美在线" in joined or "盖成绩" in joined:
            return row
    return sheet.max_row + 1


def copy_row_style(sheet, source_row, target_row):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)


def clear_row_values(sheet, row, max_column=BASE_COLUMN_COUNT):
    for column in range(1, max_column + 1):
        sheet.cell(row, column).value = None
        sheet.cell(row, column).hyperlink = None


def is_yellow_fill(fill):
    return fill and fill.fill_type == "solid" and fill.fgColor.rgb == YELLOW_FILL_RGB


def remove_yellow_fills(sheet):
    from openpyxl.styles import PatternFill

    no_fill = PatternFill()
    for row in sheet.iter_rows():
        for cell in row:
            if is_yellow_fill(cell.fill):
                cell.fill = copy.copy(no_fill)


def first_cell_with_fill(sheet, rgb):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == rgb:
                return cell
    return None


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy.copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy.copy(source.alignment)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)


def apply_border(sheet, min_row, max_row, min_column=1, max_column=BASE_COLUMN_COUNT):
    from openpyxl.styles import Border, Side

    thin = Side(style="thin", color="FF000000")
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            sheet.cell(row, column).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )


def set_body_font(cell):
    from openpyxl.styles import Font

    cell.font = Font(
        name=BODY_FONT_NAME,
        sz=BODY_FONT_SIZE,
        bold=cell.font.bold,
        italic=cell.font.italic,
        vertAlign=cell.font.vertAlign,
        underline=cell.font.underline,
        strike=cell.font.strike,
        color=copy.copy(cell.font.color),
    )


def apply_body_font(sheet, min_row, max_row, min_column=1, max_column=BASE_COLUMN_COUNT):
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            set_body_font(sheet.cell(row, column))


def set_centered_content_alignment(cell):
    from openpyxl.styles import Alignment

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
        shrink_to_fit=False,
    )


def apply_centered_content_alignment(sheet, min_row, max_row, min_column=1, max_column=BASE_COLUMN_COUNT):
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            set_centered_content_alignment(sheet.cell(row, column))


def apply_header_style(sheet, min_row, max_row, min_column=1, max_column=BASE_COLUMN_COUNT):
    source = first_cell_with_fill(sheet, HEADER_FILL_RGB)
    if not source:
        return
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            copy_cell_style(source, sheet.cell(row, column))


def ensure_title_span(sheet, end_column=BASE_COLUMN_COUNT):
    from openpyxl.styles import Alignment

    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row == 1 and merged_range.max_row == 1:
            sheet.unmerge_cells(str(merged_range))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=cell.alignment.wrap_text,
    )


def ensure_score_detail_header(sheet, end_column=BASE_COLUMN_COUNT):
    from openpyxl.styles import Alignment

    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row == 5 and merged_range.max_row == 5:
            sheet.unmerge_cells(str(merged_range))
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=end_column)
    cell = sheet.cell(5, 1)
    cell.value = "成绩明细"
    cell.alignment = copy.copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=cell.alignment.wrap_text,
    )
    apply_header_style(sheet, 5, 5)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=cell.alignment.wrap_text,
    )


def ensure_top_header_layout(sheet):
    from openpyxl.styles import Alignment

    procurement_contact = sheet["D4"].value
    project_manager = sheet["E4"].value
    assessment_manager = sheet["F4"].value
    for merged_range in list(sheet.merged_cells.ranges):
        if 2 <= merged_range.min_row <= 4 or 2 <= merged_range.max_row <= 4:
            sheet.unmerge_cells(str(merged_range))
    for range_ref in TOP_HEADER_MERGES:
        sheet.merge_cells(range_ref)
    apply_header_style(sheet, 2, 3)
    for column_letter, width in TOP_HEADER_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column_letter].width = width

    labels = {
        "A2": "单位名称",
        "B2": "考试名称",
        "D2": "考试时间",
        "F2": "采购方联系人",
        "G2": "项目经理（ATA）",
        "I2": "测评经理（ATA）",
    }
    for coordinate, value in labels.items():
        cell = sheet[coordinate]
        cell.value = value
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=cell.alignment.wrap_text,
        )
    sheet["F4"] = procurement_contact
    sheet["G4"] = project_manager
    sheet["I4"] = assessment_manager
    sheet.row_dimensions[4].height = TOP_CONTENT_ROW_HEIGHT
    for coordinate in ("A4", "B4", "D4", "F4", "G4", "I4"):
        set_centered_content_alignment(sheet[coordinate])
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def write_row(sheet, row_index, row, default_course):
    exam_status = normalized_status(row.get("exam_status"))
    score = score_value(row.get("score"))
    if exam_status == "缺考" and score == "":
        score = "--"
    values = [
        text(row.get("name")),
        text(row.get("gender")),
        text(row.get("identity_id")),
        text(row.get("mobile")),
        text(row.get("email")),
        text(row.get("course") or default_course),
        text(row.get("permit")),
        exam_status,
        score,
        text(row.get("violation") or "无"),
    ]
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column)
        cell.value = value
        if column in TEXT_COLUMNS:
            cell.number_format = "@"
        set_centered_content_alignment(cell)
    sheet.row_dimensions[row_index].height = max(
        sheet.row_dimensions[row_index].height or 0,
        CONTENT_ROW_HEIGHT,
    )


def normalized_reports(row):
    reports = row.get("reports") or row.get("score_reports") or row.get("assessment_reports") or []
    if isinstance(reports, dict):
        reports = list(reports.values())
    if not isinstance(reports, list):
        return []
    normalized = []
    for report in reports:
        if not isinstance(report, dict):
            continue
        name = text(report.get("name") or report.get("title") or report.get("label")).strip()
        url = text(report.get("url") or report.get("link") or report.get("href")).strip()
        status = text(report.get("status")).strip()
        if not name or not url:
            continue
        normalized.append({"name": name, "url": url, "status": status})
    return normalized


def report_columns(rows):
    names = []
    seen = set()
    for row in rows:
        for report in normalized_reports(row):
            key = report["name"]
            if key in seen:
                continue
            seen.add(key)
            names.append(key)
    return names


def report_by_name(row):
    return {report["name"]: report for report in normalized_reports(row)}


def report_column_width(name):
    return max(22, min(42, len(name) + 8))


def apply_report_columns(sheet, rows, report_names, last_data_row):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if not report_names:
        return

    header_source = sheet.cell(6, BASE_COLUMN_COUNT)
    body_source = sheet.cell(DATA_START_ROW, BASE_COLUMN_COUNT)
    start_column = BASE_COLUMN_COUNT + 1
    for offset, name in enumerate(report_names):
        column = start_column + offset
        column_letter = get_column_letter(column)
        sheet.column_dimensions[column_letter].width = report_column_width(name)

        header_cell = sheet.cell(6, column)
        copy_cell_style(header_source, header_cell)
        header_cell.value = name
        set_centered_content_alignment(header_cell)

        for row_index in range(DATA_START_ROW, last_data_row + 1):
            body_cell = sheet.cell(row_index, column)
            copy_cell_style(body_source, body_cell)
            set_centered_content_alignment(body_cell)
            body_cell.value = None
            body_cell.hyperlink = None

    for offset, row in enumerate(rows):
        row_reports = report_by_name(row)
        row_index = DATA_START_ROW + offset
        for report_name, report in row_reports.items():
            if report_name not in report_names:
                continue
            column = start_column + report_names.index(report_name)
            cell = sheet.cell(row_index, column)
            cell.value = report["name"]
            cell.hyperlink = report["url"]
            cell.font = Font(
                name=BODY_FONT_NAME,
                sz=BODY_FONT_SIZE,
                color="FF0563C1",
                underline="single",
            )

    apply_border(sheet, 6, last_data_row, start_column, start_column + len(report_names) - 1)


def export_score_feedback(template_path, payload_path, output_path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，请先安装依赖") from exc

    template_path = Path(template_path)
    payload_path = Path(payload_path)
    output_path = Path(output_path)
    if not template_path.exists():
        return {"ok": False, "errors": [f"成绩单模板不存在：{template_path}"]}

    payload = json.loads(payload_path.read_text("utf-8"))
    rows = payload.get("rows") or []
    report_names = report_columns(rows)
    content_column_count = BASE_COLUMN_COUNT + len(report_names)
    workbook = load_workbook(template_path)
    sheet = workbook.active

    exam_name = text(payload.get("examName") or "")
    exam_time = text(payload.get("examTime") or "")
    today = date.today()
    processed_date = text(payload.get("processedDate") or f"{today.year}年{today.month}月{today.day}日")
    ensure_title_span(sheet, content_column_count)
    ensure_top_header_layout(sheet)
    sheet["B4"] = exam_name
    sheet["D4"] = exam_time
    apply_border(sheet, 4, 4)
    apply_body_font(sheet, 4, 4)

    footer_row = find_footer_row(sheet)
    footer_gap_row = footer_row - 1
    existing_slots = max(0, footer_gap_row - DATA_START_ROW)
    required_rows = max(1, len(rows))
    if required_rows > existing_slots:
        insert_count = required_rows - existing_slots
        sheet.insert_rows(footer_gap_row, insert_count)
        footer_row += insert_count
        footer_gap_row += insert_count

    footer_date_cell = sheet.cell(footer_row + 1, 6)
    footer_date_cell.value = processed_date
    footer_date_cell.number_format = "@"

    last_data_row = footer_gap_row - 1
    for row_index in range(DATA_START_ROW, last_data_row + 1):
        copy_row_style(sheet, DATA_START_ROW, row_index)
        clear_row_values(sheet, row_index, content_column_count)
    clear_row_values(sheet, footer_gap_row, content_column_count)

    default_course = exam_name
    for offset, row in enumerate(rows):
        write_row(sheet, DATA_START_ROW + offset, row, default_course)
    if rows:
        apply_body_font(sheet, DATA_START_ROW, DATA_START_ROW + len(rows) - 1)

    for column in TEXT_COLUMNS:
        for row_index in range(DATA_START_ROW, max(DATA_START_ROW, DATA_START_ROW + len(rows))):
            sheet.cell(row_index, column).number_format = "@"

    remove_yellow_fills(sheet)
    ensure_score_detail_header(sheet, content_column_count)
    apply_report_columns(sheet, rows, report_names, last_data_row)
    apply_centered_content_alignment(sheet, 6, 6, 1, content_column_count)
    last_content_row = max(6, DATA_START_ROW + len(rows) - 1)
    apply_body_font(sheet, 2, last_content_row, 1, content_column_count)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "ok": True,
        "path": str(output_path),
        "rows": len(rows),
        "reportColumns": len(report_names),
        "errors": [],
    }


def main():
    if len(sys.argv) != 4:
        print(json.dumps({"ok": False, "errors": ["用法：score_feedback_exporter.py template.xlsx payload.json output.xlsx"]}, ensure_ascii=False))
        return 1
    try:
        result = export_score_feedback(sys.argv[1], sys.argv[2], sys.argv[3])
        print(json.dumps(result, ensure_ascii=False))
        return 0 if result.get("ok") else 2
    except Exception as exc:
        print(json.dumps({"ok": False, "errors": [str(exc)]}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
